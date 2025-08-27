"""
Extract cleaned OMC data from organized Excel files
Following the same pattern as BDC extraction
"""

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class OMCDataExtractor:
    def __init__(self):
        self.base_path = r'C:\Users\victo\Documents\Data_Science_Projects\petroverse_analytics\data\raw\Raw_Organized_OMC'
        self.month_mapping = {
            'jan': 1, 'january': 1,
            'feb': 2, 'february': 2,
            'mar': 3, 'march': 3,
            'apr': 4, 'april': 4,
            'may': 5,
            'jun': 6, 'june': 6,
            'jul': 7, 'july': 7,
            'aug': 8, 'august': 8,
            'sep': 9, 'september': 9, 'sept': 9,
            'oct': 10, 'october': 10,
            'nov': 11, 'november': 11,
            'dec': 12, 'december': 12
        }
        self.all_data = []
        
    def analyze_file_structure(self):
        """Analyze OMC file structure to understand format"""
        logger.info("\nANALYZING OMC FILE STRUCTURE:")
        logger.info("=" * 60)
        
        files = [f for f in os.listdir(self.base_path) if f.endswith('.xlsx')]
        
        for file_name in files[:2]:  # Check first 2 files
            file_path = os.path.join(self.base_path, file_name)
            logger.info(f"\nFile: {file_name}")
            
            try:
                # Load workbook to get sheet names
                wb = load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                logger.info(f"  Sheets: {sheets[:5]}...")  # Show first 5 sheets
                
                # Check a sample sheet structure
                for sheet_name in sheets[:2]:
                    logger.info(f"\n  Checking sheet: {sheet_name}")
                    df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                    
                    # Check for header row
                    if not df.empty:
                        logger.info(f"    Columns (row 0): {list(df.columns)[:5]}")
                        
                        # Check with header at row 1
                        df_h1 = pd.read_excel(file_path, sheet_name=sheet_name, header=1, nrows=5)
                        if not df_h1.empty:
                            logger.info(f"    Columns (header=1): {list(df_h1.columns)[:8]}")
                
                wb.close()
                
            except Exception as e:
                logger.error(f"  Error analyzing file: {e}")
    
    def extract_sheet_data(self, file_path, sheet_name, year):
        """Extract data from a single sheet"""
        try:
            # Try reading with header at row 1 (like BDC files)
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
            
            # Get the month from sheet name
            month_str = sheet_name.lower().strip()
            month = self.month_mapping.get(month_str, None)
            
            if month is None:
                # Try extracting month from sheet name variations
                for key in self.month_mapping:
                    if key in month_str:
                        month = self.month_mapping[key]
                        break
            
            if month is None:
                logger.warning(f"    Could not determine month from sheet name: {sheet_name}")
                return 0
            
            # Find the company column (first column usually)
            company_col = None
            for col in df.columns:
                if isinstance(col, str):
                    col_lower = col.lower()
                    if 'company' in col_lower or 'omc' in col_lower or col_lower == 'unnamed: 0':
                        company_col = col
                        break
            
            if company_col is None and len(df.columns) > 0:
                # Use first column as company column
                company_col = df.columns[0]
            
            if company_col is None:
                logger.warning(f"    No company column found in {sheet_name}")
                return 0
            
            # Process each row (company)
            records_added = 0
            for idx, row in df.iterrows():
                company_name = row[company_col]
                
                # Skip empty company names or total rows
                if pd.isna(company_name) or company_name == '' or 'total' in str(company_name).lower():
                    continue
                
                # Clean company name
                company_name = str(company_name).strip()
                
                # Process each product column
                for col in df.columns:
                    if col == company_col:
                        continue
                    
                    # Skip unnamed or empty columns
                    if pd.isna(col) or 'unnamed' in str(col).lower():
                        continue
                    
                    product = str(col).strip()
                    volume = row[col]
                    
                    # Skip empty or zero volumes
                    if pd.isna(volume) or volume == 0:
                        continue
                    
                    # Create record
                    record = {
                        'source_file': os.path.basename(file_path),
                        'sheet_name': sheet_name,
                        'extraction_date': datetime.now().date(),
                        'year': year,
                        'month': month,
                        'period_date': f"{year}-{month:02d}-01",
                        'period_type': 'monthly',
                        'company_name': company_name,
                        'product_code': '',
                        'product_original_name': product,
                        'product': product,  # Will be standardized later
                        'unit_type': 'LITERS',  # Default, will check for LPG
                        'volume': float(volume),
                        'volume_liters': float(volume),
                        'volume_kg': 0,  # Will calculate
                        'volume_mt': 0,  # Will calculate
                        'company_type': 'OMC'
                    }
                    
                    self.all_data.append(record)
                    records_added += 1
            
            return records_added
            
        except Exception as e:
            logger.error(f"    Error extracting sheet {sheet_name}: {e}")
            return 0
    
    def extract_all_data(self):
        """Extract data from all OMC files"""
        logger.info("\nEXTRACTING OMC DATA:")
        logger.info("=" * 60)
        
        files = sorted([f for f in os.listdir(self.base_path) if f.endswith('.xlsx')])
        
        for file_name in files:
            # Extract year from filename
            year = None
            for y in range(2019, 2026):
                if str(y) in file_name:
                    year = y
                    break
            
            if year is None:
                logger.warning(f"Could not determine year from filename: {file_name}")
                continue
            
            logger.info(f"\nProcessing: {file_name} (Year: {year})")
            file_path = os.path.join(self.base_path, file_name)
            
            try:
                # Get all sheet names
                wb = load_workbook(file_path, read_only=True, data_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                # Filter for month sheets
                month_sheets = []
                for sheet in sheets:
                    sheet_lower = sheet.lower().strip()
                    for month_name in self.month_mapping.keys():
                        if month_name in sheet_lower:
                            month_sheets.append(sheet)
                            break
                
                logger.info(f"  Found {len(month_sheets)} month sheets")
                
                # Extract data from each month sheet
                for sheet_name in month_sheets:
                    records = self.extract_sheet_data(file_path, sheet_name, year)
                    if records > 0:
                        logger.info(f"    {sheet_name}: {records} records extracted")
                
            except Exception as e:
                logger.error(f"  Error processing file {file_name}: {e}")
        
        logger.info(f"\nTotal OMC records extracted: {len(self.all_data):,}")
    
    def save_raw_data(self):
        """Save extracted raw OMC data"""
        if not self.all_data:
            logger.warning("No data to save")
            return None
        
        df = pd.DataFrame(self.all_data)
        
        # Save raw extracted data
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'CLEANED_OMC_data_{timestamp}.csv'
        df.to_csv(output_file, index=False)
        
        logger.info(f"\nRaw OMC data saved to: {output_file}")
        
        # Summary statistics
        logger.info("\nEXTRACTION SUMMARY:")
        logger.info("=" * 50)
        logger.info(f"Total records: {len(df):,}")
        logger.info(f"Years covered: {df['year'].min()} - {df['year'].max()}")
        logger.info(f"Unique companies: {df['company_name'].nunique()}")
        logger.info(f"Unique products: {df['product'].nunique()}")
        
        # Sample of products found
        logger.info("\nSample products found:")
        for product in df['product'].value_counts().head(10).index:
            count = df[df['product'] == product].shape[0]
            logger.info(f"  {product}: {count:,} records")
        
        # Sample of companies found
        logger.info("\nTop 10 companies by record count:")
        for company in df['company_name'].value_counts().head(10).index:
            count = df[df['company_name'] == company].shape[0]
            logger.info(f"  {company}: {count:,} records")
        
        return output_file, df

def main():
    extractor = OMCDataExtractor()
    
    # First analyze structure
    extractor.analyze_file_structure()
    
    # Extract all data
    extractor.extract_all_data()
    
    # Save raw data
    output_file, df = extractor.save_raw_data()
    
    return output_file, df

if __name__ == "__main__":
    print("OMC DATA EXTRACTION FROM CLEANED FILES")
    print("=" * 60)
    output_file, df = main()
    if output_file:
        print(f"\nExtraction complete! Data saved to: {output_file}")
    else:
        print("\nNo data extracted.")